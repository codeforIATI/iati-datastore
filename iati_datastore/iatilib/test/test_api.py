import os
import json
from datetime import datetime
from xml.etree import ElementTree as ET
import csv
from io import StringIO
from tempfile import mkstemp

import mock
import openpyxl

from . import factories as fac
from . import ClientTestCase
from iatilib import parse, db, model

from iatilib.currency_conversion import update_exchange_rates

def read_fixture(fix_name, encoding='utf-8'):
    """Read and convert fixture from csv file"""
    return csv.reader(open(fixture_filename(fix_name)).read().strip().split("\n"), delimiter=',')

class TestLatestApiRedirect(ClientTestCase):
    def test_latest_api_redirect(self):
        resp = self.client.get('/api/')
        self.assertEquals(302, resp.status_code)
        self.assertRegex(resp.headers['Location'], '/api/1/$')


class TestListRoutes(ClientTestCase):
    def test_list_routes(self):
        resp = self.client.get('/api/1/')
        self.assertEquals(200, resp.status_code)
        self.assertEquals("application/json", resp.content_type)
        data = json.loads(resp.data)
        self.assertIn('http://localhost/api/1/', data)


class TestAbout(ClientTestCase):
    def test_about_http(self):
        resp = self.client.get('/api/1/about/')
        self.assertEquals(200, resp.status_code)


class TestAboutDatasets(ClientTestCase):
    def test_about(self):
        fac.DatasetFactory.create(
            name='tst-old',
            resources=[fac.ResourceFactory.create(
                url="http://foo",
            )]
        )
        resp = self.client.get('/api/1/about/dataset/')
        data = json.loads(resp.data)
        self.assertEquals(200, resp.status_code)
        self.assertIn("datasets", data)
        self.assertEquals("tst-old", data["datasets"][0])

    def test_about_details(self):
        fac.DatasetFactory.create(
            name='tst-old',
            resources=[fac.ResourceFactory.create(
                url="http://foo",
            )]
        )
        resp = self.client.get('/api/1/about/dataset/?detail=true')
        data = json.loads(resp.data)
        self.assertEquals(200, resp.status_code)
        self.assertIn("datasets", data)
        self.assertEquals('http://foo', data["datasets"][0]['url'])
        self.assertIn("ok", data)
        self.assertIn("total-count", data)

    def test_about_dataset(self):
        fac.DatasetFactory.create(
            name='tst-old',
            resources=[fac.ResourceFactory.create(
                url="http://foo",
            )]
        )
        resp = self.client.get('/api/1/about/dataset/tst-old/')
        data = json.loads(resp.data)
        self.assertEquals(200, resp.status_code)
        self.assertEquals('tst-old', data["dataset"])
        self.assertEquals(1, data["num_resources"])
        self.assertEquals('http://foo', data["resources"][0]['url'])

    def test_about_not_a_dataset(self):
        resp = self.client.get('/api/1/about/dataset/bad-dataset-id/')
        self.assertEquals(404, resp.status_code)

    def test_about_invalid_filter(self):
        resp = self.client.get('/api/1/about/dataset/?invalid=true')
        self.assertEquals(400, resp.status_code)
        self.assertIn(
            "Invalid arguments passed as filter",
            resp.data.decode())


class TestErrorDatasets(ClientTestCase):
    def test_errors(self):
        fac.LogFactory.create()
        resp = self.client.get('/api/1/error/dataset/')
        data = json.loads(resp.data)
        self.assertEquals(200, resp.status_code)
        self.assertIn("errored_datasets", data)
        self.assertEquals(
            'bad-dataset',
            data['errored_datasets'][0]['dataset'])

    def test_errors_log(self):
        fac.LogFactory.create()
        resp = self.client.get('/api/1/error/dataset.log')
        self.assertEquals(200, resp.status_code)
        self.assertEquals(
            "text/plain; charset=utf-8", resp.content_type)

    def test_error_log(self):
        fac.LogFactory.create()
        resp = self.client.get('/api/1/error/dataset.log/bad-dataset/')
        self.assertEquals(200, resp.status_code)
        self.assertEquals(
            "text/plain; charset=utf-8", resp.content_type)

    def test_error(self):
        fac.LogFactory.create()
        resp = self.client.get('/api/1/error/dataset/bad-dataset/')
        data = json.loads(resp.data)
        self.assertEquals(200, resp.status_code)
        self.assertIn("errors", data)
        self.assertEquals(
            'bad-dataset',
            data['errors'][0]['dataset'])
        self.assertEquals(
            'Dataset is broken',
            data['errors'][0]['msg'])


class TestDeletedActivitiesView(ClientTestCase):
    def test_deleted_activities(self):
        db.session.add(model.DeletedActivity(
            iati_identifier='test',
            deletion_date=datetime(2000, 1, 1))
        )
        db.session.commit()
        resp = self.client.get('api/1/about/deleted/')
        data = json.loads(resp.data)
        deleted_activities = data['deleted_activities']
        self.assertEquals("test", deleted_activities[0]['iati_identifier'])
        self.assertEquals("2000-01-01", deleted_activities[0]['deletion_date'])

    def test_deleted_invalid_filter(self):
        resp = self.client.get('/api/1/about/deleted/?invalid=true')
        self.assertEquals(400, resp.status_code)
        self.assertIn(
            "Invalid arguments passed as filter",
            resp.data.decode())


class TestMetaFilters(ClientTestCase):
    def test_http(self):
        resp = self.client.get('/api/1/meta/filters/')
        self.assertEquals(200, resp.status_code)
        self.assertEquals("application/json", resp.content_type)
        self.assertIn("filters", json.loads(resp.data))


class TestEmptyDb_JSON(ClientTestCase):
    url = '/api/1/access/activity/'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("application/json", resp.content_type)

    def tests_json_decode(self):
        resp = self.client.get(self.url)
        self.assert_(json.loads(resp.data))

    def test_json_ok(self):
        resp = self.client.get(self.url)
        js = json.loads(resp.data)
        self.assertTrue(js["ok"])

    def test_json_results(self):
        resp = self.client.get(self.url)
        js = json.loads(resp.data)
        self.assertEquals(js["iati-activities"], [])


class TestEmptyDb_XML(ClientTestCase):
    """
    Raw XML for empty db.

    Basic layout (see: https://github.com/okfn/iati-datastore/issues/14)
    <result>
       <page>5</page>
       ...metadata here...
       <result-activities>...concatted string...
       </result-activities>
    </result>
    """
    url = '/api/1/access/activity.xml'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("application/xml; charset=utf-8", resp.content_type)

    def test_decode(self):
        resp = self.client.get(self.url)
        # an ElementTree node object does not test as true
        self.assert_(hasattr(ET.fromstring(resp.get_data(as_text=True)), "tag"))

    def test_resp_ok(self):
        resp = self.client.get(self.url)
        xml = ET.fromstring(resp.get_data(as_text=True))
        self.assertTrue(xml.find('ok').text == 'True')

    def test_results(self):
        resp = self.client.get(self.url)
        xml = ET.fromstring(resp.get_data(as_text=True))
        self.assertEquals(xml.findall('result-activities'), [])

    def test_root_element(self):
        resp = self.client.get(self.url)
        xml = ET.fromstring(resp.get_data(as_text=True))
        self.assertEquals(xml.tag, "result")


class TestEmptyDb_ActivityCSV(ClientTestCase):
    """
    CSV for empty db
    """
    url = '/api/1/access/activity.csv'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("text/csv; charset=utf-8", resp.content_type)

    def test_fields(self):
        resp = self.client.get(self.url)
        headers = next(csv.reader(StringIO(resp.get_data(as_text=True))))
        for exp in ["start-planned", "start-actual"]:
            self.assertIn(exp, headers)


class TestEmptyDb_ActivityXLSX(ClientTestCase):
    """
    XLSX for empty db
    """
    url = '/api/1/access/activity.xlsx'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resp.content_type)


class TestEmptyDb_TransactionCSV(ClientTestCase):
    """
    CSV for empty db
    """
    url = '/api/1/access/transaction.csv'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("text/csv; charset=utf-8", resp.content_type)

class TestEmptyDb_TransactionXLSX(ClientTestCase):
    """
    XLSX for empty db
    """
    url = '/api/1/access/transaction.xlsx'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resp.content_type)


class TestEmptyDb_BudgetCSV(ClientTestCase):
    """
    CSV for empty db
    """
    url = '/api/1/access/budget.csv'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("text/csv; charset=utf-8", resp.content_type)


class TestEmptyDb_BudgetXLSX(ClientTestCase):
    """
    XLSX for empty db
    """
    url = '/api/1/access/budget.xlsx'

    def test_http_ok(self):
        resp = self.client.get(self.url)
        self.assertEquals(200, resp.status_code)

    def test_content_type(self):
        resp = self.client.get(self.url)
        self.assertEquals("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resp.content_type)


def fixture_filename(fix_name):
    return os.path.join(
            os.path.dirname(__file__), "fixtures", fix_name)


def load_fix(fix_name):
    activities = parse.document_from_file(fixture_filename(fix_name))
    db.session.add_all(activities)
    db.session.commit()


class TestSingleActivity(ClientTestCase):
    """
    Different reprisentations of the same input activity
    """

    def test_xml_activity_count(self):
        load_fix("single_activity.xml")
        resp = self.client.get('/api/1/access/activity.xml')
        xml = ET.fromstring(resp.get_data(as_text=True))
        self.assertEquals(1, len(xml.findall('.//iati-activity')))

    def test_xml_activity_data(self):
        load_fix("single_activity.xml")
        in_xml = ET.parse(fixture_filename("single_activity.xml"))
        resp = self.client.get('/api/1/access/activity.xml')
        xml = ET.fromstring(resp.get_data(as_text=True))
        x1 = in_xml.find('.//iati-activity')
        x2 = xml.find('.//iati-activity')
        self.assertEquals(x1, x2)

    def test_csv_activity_count(self):
        load_fix("single_activity.xml")
        with self.client as client:
            resp = client.get('/api/1/access/activity.csv')
            self.assertEquals(2, resp.get_data(as_text=True).count("\n"))

    def test_xlsx_activity_count(self):
        load_fix("single_activity.xml")
        resp = self.client.get('/api/1/access/activity.xlsx')
        # write response to temporary file
        (handler, filename) = mkstemp(prefix="iati-datastore-classic-tests-", suffix='.xlsx')
        os.write(handler, resp.get_data())
        os.close(handler)
        # open workbook
        workbook = openpyxl.load_workbook(filename=filename)
        # check it
        worksheet = workbook.active
        self.assertEquals(2, worksheet.max_row)
        # cleanup
        workbook.close()
        os.remove(filename)

    def test_xlsx_activity_data(self):
        load_fix("single_activity.xml")
        resp = self.client.get('/api/1/access/activity.xlsx')
        # write response to temporary file
        (handler, filename) = mkstemp(prefix="iati-datastore-classic-tests-", suffix='.xlsx')
        os.write(handler, resp.get_data())
        os.close(handler)
        # open workbook
        workbook = openpyxl.load_workbook(filename=filename)
        # open sample data
        in_xml = ET.parse(fixture_filename("single_activity.xml"))
        id = in_xml.find('.//iati-activity/iati-identifier').text
        # check it
        worksheet = workbook.active
        self.assertEquals(id, worksheet['A2'].value)
        # cleanup
        workbook.close()
        os.remove(filename)

class TestManyActivities(ClientTestCase):
    def test_xml_activity_count(self):
        load_fix("many_activities.xml")
        resp = self.client.get('/api/1/access/activity.xml')
        xml = ET.fromstring(resp.get_data(as_text=True))
        self.assertEquals(2, len(xml.findall('.//iati-activity')))

    def test_xml_activity_data(self):
        load_fix("many_activities.xml")
        in_xml = ET.parse(fixture_filename("many_activities.xml"))
        resp = self.client.get('/api/1/access/activity.xml')
        xml = ET.fromstring(resp.get_data(as_text=True))
        self.assertEquals(
            ET.tostring(in_xml.find('.//iati-activity')),
            ET.tostring(xml.find('.//iati-activity')))

    def test_csv_activity_count(self):
        load_fix("many_activities.xml")
        with self.client as client:
            resp = client.get('/api/1/access/activity.csv')
            reader = csv.DictReader(StringIO(resp.get_data(as_text=True)))
            self.assertEquals(2, len(list(reader)))

    def test_xlsx_activity_count(self):
        load_fix("many_activities.xml")
        resp = self.client.get('/api/1/access/activity.xlsx')
        # write response to temporary file
        (handler, filename) = mkstemp(prefix="iati-datastore-classic-tests-", suffix='.xlsx')
        os.write(handler, resp.get_data())
        os.close(handler)
        # open workbook
        workbook = openpyxl.load_workbook(filename=filename)
        # check it
        worksheet = workbook.active
        self.assertEquals(3, worksheet.max_row)
        # cleanup
        workbook.close()
        os.remove(filename)


class TestPagination(ClientTestCase):
    def test_invalid_page(self):
        resp = self.client.get('/api/1/access/activity/?offset=-1')
        self.assertEquals(400, resp.status_code)


class ApiViewMixin(object):
    @mock.patch('iatilib.frontend.api1.validators.activity_api_args')
    def test_validator_called(self, mock):
        self.client.get(self.base_url)
        self.assertEquals(1, mock.call_count)

    def test_filter_called(self):
        with mock.patch(self.filter) as mm:
            self.client.get(self.base_url + '?recipient-country=MW')
            self.assertEquals(1, mm.call_count)

    def test_serializer_called(self):
        with mock.patch(self.serializer) as mm:
            self.client.get(self.base_url)
            self.assertEquals(1, mm.call_count)

    def test_invalid_format(self):
        resp = self.client.get(self.base_url + ".zzz")
        self.assertEquals(404, resp.status_code)

    def test_junk_before_format(self):
        url = self.base_url[:-4] + '-bad.' + self.extension
        resp = self.client.get(url)
        self.assertEquals(404, resp.status_code)

    def test_junk_in_format(self):
        url = self.base_url[:-4] + '.bad-' + self.extension
        resp = self.client.get(url)
        self.assertEquals(404, resp.status_code)


class TestActivityView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/activity.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.ActivityCSVView.filter'
    serializer = 'iatilib.frontend.api1.ActivityCSVView.serializer'


class TestActivityXLSXView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/activity.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.ActivityXLSXView.filter'
    serializer = 'iatilib.frontend.api1.ActivityXLSXView.serializer'


class TestActivityBySectorView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/activity/by_sector.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.ActivityBySectorView.filter'
    serializer = 'iatilib.frontend.api1.ActivityBySectorView.serializer'


class TestActivityBySectorXLSXView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/activity/by_sector.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.ActivityBySectorXLSXView.filter'
    serializer = 'iatilib.frontend.api1.ActivityBySectorXLSXView.serializer'


class TestActivityByCountryView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/activity/by_country.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.ActivityByCountryView.filter'
    serializer = 'iatilib.frontend.api1.ActivityByCountryView.serializer'


class TestActivityByCountryXLSXView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/activity/by_country.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.ActivityByCountryXLSXView.filter'
    serializer = 'iatilib.frontend.api1.ActivityByCountryXLSXView.serializer'


class CommonTransactionTests(object):
    def test_reporting_org(self):
        load_fix("transaction_ref.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('reporting-org-ref')
        self.assertEquals(u'GB-CHC-285776', output[1][i])

    def test_ref_output(self):
        load_fix("transaction_ref.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_ref')
        self.assertEquals(u'36258', output[1][i])
        self.assertEquals(u'', output[2][i])

    def test_transaction_value_currency(self):
        load_fix("transaction_provider.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_value_currency')
        self.assertEquals(u'GBP', output[1][i])

    def test_transaction_value_value_date(self):
        load_fix("transaction_provider.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_value_value-date')
        self.assertEquals(u'2011-08-19', output[1][i])

    def test_provider_org_ref_output(self):
        load_fix("transaction_provider.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_provider-org_ref')
        self.assertEquals(u'GB-1-201242-101', output[1][i])

    def test_provider_org_output(self):
        load_fix("transaction_provider.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_provider-org')
        self.assertEquals(u'DFID', output[1][i])

    def test_provider_org_activity_id_output(self):
        load_fix("provider-activity-id.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_provider-org_provider-activity-id')
        self.assertEquals(u'GB-1-202907', output[1][i])

    def test_receiver_org_ref_output(self):
        load_fix("transaction_provider.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_receiver-org_ref')
        self.assertEquals(u'GB-CHC-313139', output[1][i])

    def test_receiver_org_output(self):
        """receiver_org should be in transaction.csv output"""
        load_fix("provider-activity-id.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_receiver-org')
        self.assertEquals(u'Bond', output[1][i])

    def test_receiver_org_activity_id_output(self):
        load_fix("provider-activity-id.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_receiver-org_receiver-activity-id')
        self.assertEquals(u'GB-CHC-1068839-dfid_ag_11-13', output[1][i])

    def test_description(self):
        load_fix("transaction_provider.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_description')
        self.assertEquals(
                u'Funds received from DFID for activities in Aug- Sept 2011',
                output[1][i]
        )

    def test_flow_type(self):
        load_fix("transaction_fields_code_lists.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_flow-type_code')
        self.assertEquals(u'30', output[1][i])

    def test_finance_type(self):
        load_fix("transaction_fields_code_lists.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_finance-type_code')
        self.assertEquals(u'110', output[1][i])

    def test_aid_type(self):
        load_fix("transaction_fields_code_lists.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_aid-type_code')
        self.assertEquals(u'B01', output[1][i])

    def test_tied_status(self):
        load_fix("transaction_fields_code_lists.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_tied-status_code')
        self.assertEquals(u'5', output[1][i])

    def test_disbursement_channel_status(self):
        load_fix("transaction_fields_code_lists.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('transaction_disbursement-channel_code')
        self.assertEquals(u'2', output[1][i])


class CommonTransactionXLSXTests(object):
    def _test(
        self, load_fixture_filename, row_header,
        asserts_equals_by_row={}, assert_none_rows=[], asserts_equals_date_by_row={}
    ):
        load_fix(load_fixture_filename)
        resp = self.client.get(self.base_url)
        # write response to temporary file
        (handler, filename) = mkstemp(prefix="iati-datastore-classic-tests-", suffix='.xlsx')
        os.write(handler, resp.get_data())
        os.close(handler)
        # open workbook
        workbook = openpyxl.load_workbook(filename=filename)
        # find column index
        worksheet = workbook.active
        header_row = worksheet[1]
        # (this value will be zero based, when we look up columns later it is 1-based, so +1)
        column = [i.value for i in header_row].index(row_header) + 1
        # asserts
        for row, value in asserts_equals_by_row.items():
            self.assertEquals(
                value,
                worksheet.cell(row, column).value
            )
        for row in assert_none_rows:
            self.assertEquals(
                None,
                worksheet.cell(row, column).value
            )
        for row, value in asserts_equals_date_by_row.items():
            self.assertEquals(
                value,
                str(worksheet.cell(row, column).value.date())
            )
        # cleanup
        workbook.close()
        os.remove(filename)

    def test_reporting_org(self):
        self._test(
            "transaction_ref.xml",
            'reporting-org-ref',
            asserts_equals_by_row={2: u'GB-CHC-285776'}
        )

    def test_ref_output(self):
        self._test(
            "transaction_ref.xml",
            'transaction_ref',
            asserts_equals_by_row={2: u'36258'},
            assert_none_rows=[3]
        )

    def test_transaction_value_currency(self):
        self._test(
            "transaction_provider.xml",
            'transaction_value_currency',
            asserts_equals_by_row={2: u'GBP'}
        )

    def test_transaction_value_value_date(self):
        self._test(
            "transaction_provider.xml",
            'transaction_value_value-date',
            asserts_equals_date_by_row={2: u'2011-08-19'}
        )

    def test_provider_org_ref_output(self):
        self._test(
            "transaction_provider.xml",
            'transaction_provider-org_ref',
            asserts_equals_by_row={2: u'GB-1-201242-101'}
        )

    def test_provider_org_output(self):
        self._test(
            "transaction_provider.xml",
            'transaction_provider-org',
            asserts_equals_by_row={2: u'DFID'}
        )

    def test_provider_org_activity_id_output(self):
        self._test(
            "provider-activity-id.xml",
            'transaction_provider-org_provider-activity-id',
            asserts_equals_by_row={2: u'GB-1-202907'}
        )

    def test_receiver_org_ref_output(self):
        self._test(
            "transaction_provider.xml",
            'transaction_receiver-org_ref',
            asserts_equals_by_row={2: u'GB-CHC-313139'}
        )

    def test_receiver_org_output(self):
        self._test(
            "provider-activity-id.xml",
            'transaction_receiver-org',
            asserts_equals_by_row={2: u'Bond'}
        )

    def test_receiver_org_activity_id_output(self):
        self._test(
            "provider-activity-id.xml",
            'transaction_receiver-org_receiver-activity-id',
            asserts_equals_by_row={2: u'GB-CHC-1068839-dfid_ag_11-13'}
        )

    def test_description(self):
        self._test(
            "transaction_provider.xml",
            'transaction_description',
            asserts_equals_by_row={2: u'Funds received from DFID for activities in Aug- Sept 2011'}
        )

    def test_flow_type(self):
        self._test(
            "transaction_fields_code_lists.xml",
            'transaction_flow-type_code',
            asserts_equals_by_row={2: u'30'}
        )

    def test_finance_type(self):
        self._test(
            "transaction_fields_code_lists.xml",
            'transaction_finance-type_code',
            asserts_equals_by_row={2: u'110'}
        )

    def test_aid_type(self):
        self._test(
            "transaction_fields_code_lists.xml",
            'transaction_aid-type_code',
            asserts_equals_by_row={2: u'B01'}
        )

    def test_tied_status(self):
        self._test(
            "transaction_fields_code_lists.xml",
            'transaction_tied-status_code',
            asserts_equals_by_row={2: u'5'}
        )

    def test_disbursement_channel_status(self):
        self._test(
            "transaction_fields_code_lists.xml",
            'transaction_disbursement-channel_code',
            asserts_equals_by_row={2: u'2'}
        )


class TestTransactionView(ClientTestCase, ApiViewMixin, CommonTransactionTests):
    base_url = '/api/1/access/transaction.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.TransactionsView.filter'
    serializer = 'iatilib.frontend.api1.TransactionsView.serializer'


class TestTransactionXLSXView(ClientTestCase, ApiViewMixin, CommonTransactionXLSXTests):
    base_url = '/api/1/access/transaction.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.TransactionsXLSXView.filter'
    serializer = 'iatilib.frontend.api1.TransactionsXLSXView.serializer'


class TestTransactionByCountryView(ClientTestCase, ApiViewMixin, CommonTransactionTests):
    base_url = '/api/1/access/transaction/by_country.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.TransactionsByCountryView.filter'
    serializer = 'iatilib.frontend.api1.TransactionsByCountryView.serializer'


class TestTransactionByCountryXLSXView(ClientTestCase, ApiViewMixin, CommonTransactionXLSXTests):
    base_url = '/api/1/access/transaction/by_country.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.TransactionsByCountryXLSXView.filter'
    serializer = 'iatilib.frontend.api1.TransactionsByCountryXLSXView.serializer'


class TestTransactionBySectorView(ClientTestCase, ApiViewMixin, CommonTransactionTests):
    base_url = '/api/1/access/transaction/by_sector.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.TransactionsBySectorView.filter'
    serializer = 'iatilib.frontend.api1.TransactionsBySectorView.serializer'


class TestTransactionBySectorXLSXView(ClientTestCase, ApiViewMixin, CommonTransactionXLSXTests):
    base_url = '/api/1/access/transaction/by_sector.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.TransactionsBySectorXLSXView.filter'
    serializer = 'iatilib.frontend.api1.TransactionsBySectorXLSXView.serializer'


class TestBudgetView(ClientTestCase):
    base_url = '/api/1/access/budget.csv'
    filter = 'iatilib.frontend.api1.dsfilter.budgets'
    serializer = 'iatilib.frontend.api1.serialize.budget_csv'


class TestBudgetXLSXView(ClientTestCase):
    base_url = '/api/1/access/budget.xlsk'
    filter = 'iatilib.frontend.api1.dsfilter.budgets'
    serializer = 'iatilib.frontend.api1.serialize.budget_xlsx'


class TestBudgetByCountryView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/budget/by_country.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.BudgetsByCountryView.filter'
    serializer = 'iatilib.frontend.api1.BudgetsByCountryView.serializer'


class TestBudgetByCountryXLSXView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/budget/by_country.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.BudgetsByCountryXLSXView.filter'
    serializer = 'iatilib.frontend.api1.BudgetsByCountryXLSXView.serializer'


class TestBudgetBySectorView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/budget/by_sector.csv'
    extension = 'csv'
    filter = 'iatilib.frontend.api1.BudgetsBySectorView.filter'
    serializer = 'iatilib.frontend.api1.BudgetsBySectorView.serializer'


class TestBudgetBySectorXLSXView(ClientTestCase, ApiViewMixin):
    base_url = '/api/1/access/budget/by_sector.xlsx'
    extension = 'xlsx'
    filter = 'iatilib.frontend.api1.BudgetsBySectorXLSXView.filter'
    serializer = 'iatilib.frontend.api1.BudgetsBySectorXLSXView.serializer'

class TestActivityLocalesDescriptionTypes(ClientTestCase):
    """Test new functionality to output locale appropriate titles and descriptions, and new columns for each description type"""

    base_url = '/api/1/access/activity.csv'

    def test_csv_activity_count(self):
        load_fix("multiple_locales_descriptions.xml")
        resp = self.client.get(self.base_url)
        self.assertEquals(2, resp.get_data(as_text=True).count("\n"))

    def test_english_title(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('title')
        self.assertEquals(
                u'Algeria - Emergency Humanitarian Assistance to Flood Victims of the year 2001',
                output[1][i]
        )

    def test_french_title(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url + '?locale=fr').get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('title')
        self.assertEquals(
                u'Algérie - Aide Humanitaire d’Urgence aux Victimes des Inondations de 2001',
                output[1][i]
        )

    def test_english_description(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description')
        self.assertEquals(
                u"The present operation relates to emergency humanitarian aid to the victims of the 2001 floods. It’s designed because of the natural disasters to which Algeria is increasingly exposed (earthquake, drought and floods) on the one hand and, on the other hand, the lack of a national response mechanism to deal with such situations. The operation should help to mitigate the negative impact on the living conditions of populations already bruised by years of social turbulence. The Bank's assistance aims to support the action of the government and local authorities to: (i) ensure better control of the nutritional situation through the provision of foodstuffs and the provision of material resources for five (05) school canteens; (ii) ensure better control of the epidemiological situation by providing vaccines, kits, consumables, pharmaceuticals and accessories to enable health facilities to operate under acceptable conditions; (iii) supply essential vaccines and medicines for health facilities and provide equipment and teaching equipment for schools severely affected by the floods in the area of operation.",
                output[1][i]
        )

    def test_french_description(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url + '?locale=fr').get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description')
        self.assertEquals(
                u"La présente opération est relative à l’aide humanitaire d’urgence aux victimes des inondations de 2001. Elle a été conçue en raison des calamités naturelles auxquelles est de plus en plus exposée l’Algérie (tremblement de terre, sécheresse et inondations) d’une part et d’autre part, de l’absence de dispositif de riposte au niveau national pour faire face à de telles situations. L’opération devrait permettre d’atténuer les incidences négatives sur les conditions de vie de populations déjà meurtries par des années de turbulence sociale. L’assistance de la Banque vise à appuyer l’action du gouvernement et des autorités locales en vue de: i) assurer un meilleur contrôle de la situation nutritionnelle, par la fourniture de denrées alimentaires et par la dotation en moyens matériels de cinq (05) cantines scolaires; ii) assurer un meilleur contrôle de la situation épidémiologique par la provision de vaccins, kits, consommables, produits pharmaceutiques et accessoires pour permettre aux formations sanitaires de fonctionner dans des conditions acceptables ; iii) procéder à l’approvisionnement en vaccins et médicaments essentiels des formations sanitaires et à la dotation en matériels et équipements didactiques des établissements scolaires durement touchés par les inondations dans la zone d’opération.",
                output[1][i]
        )

    def test_description_general(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description')
        j = csv_headers.index('description_general')
        self.assertEquals(
                output[1][i],
                output[1][j]
        )

    def test_english_description_objectives(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description_objectives')
        self.assertEquals(
                u"The objective of the humanitarian aid is to help the populations victims of the floods of November 2001 in the zones of the disaster to prevent the epidemics (yellow fever, cholera, meningitis cerebrospinal, etc.) and to ward off the deterioration of the conditions of life.",
                output[1][i]
        )

    def test_french_description_objectives(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url + '?locale=fr').get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description_objectives')
        self.assertEquals(
                u"L’objectif de l’aide humanitaire est de secourir les populations victimes des inondations de novembre 2001 dans les zones du sinistre afin de prévenir les épidémies (fièvre jaune, choléra, méningite cérébro-spinale, etc.) et de conjurer la dégradation des conditions de vie.",
                output[1][i]
        )

    def test_english_description_target_groups(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description_target_groups')
        self.assertEquals(
                u"The beneficiaries are at-risk populations and victims housed in makeshift shelters in flooded localities. The direct beneficiaries are children and mothers who are victims of the disaster and still suffering from its consequences.",
                output[1][i]
        )

    def test_french_description_target_groups(self):
        load_fix("multiple_locales_descriptions.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url + '?locale=fr').get_data(as_text=True))))
        csv_headers = output[0]
        i = csv_headers.index('description_target_groups')
        self.assertEquals(
                u"Les bénéficiaires sont les populations à risque et victimes hébergées dans les abris de fortune dans les localités inondées. Les bénéficiaires directs en sont les enfants et les mères victimes de la catastrophe et souffrant encore de ses conséquences.",
                output[1][i]
        )

class TestActivityCurrencyConversionOutput(ClientTestCase):
    """Test new functionality to output USD and EUR"""

    base_url = '/api/1/access/activity.csv'

    def test_csv_activity_count(self):
        self.data = read_fixture("imf_exchangerates.csv")
        next(self.data, None)
        update_exchange_rates(self.data)
        load_fix("transaction-currencies.xml")
        resp = self.client.get(self.base_url)
        self.assertEquals(2, resp.get_data(as_text=True).count("\n"))

    def test_usd_currency_fields(self):
        self.data = read_fixture("imf_exchangerates.csv")
        next(self.data, None)
        update_exchange_rates(self.data)
        load_fix("transaction-currencies.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        self.assertEquals(u'12715.24',  output[1][csv_headers.index('total-Commitment-USD')])
        self.assertEquals(u'7523.84',  output[1][csv_headers.index('total-Disbursement-USD')])
        self.assertEquals(u'8821.69',  output[1][csv_headers.index('total-Expenditure-USD')])
        self.assertEquals(u'11417.39',  output[1][csv_headers.index('total-Incoming Funds-USD')])
        self.assertEquals(u'21800.19',  output[1][csv_headers.index('total-Interest Repayment-USD')])
        self.assertEquals(u'34778.69',  output[1][csv_headers.index('total-Loan Repayment-USD')])
        self.assertEquals(u'73714.19',  output[1][csv_headers.index('total-Reimbursement-USD')])

    def test_eur_currency_fields(self):
        self.data = read_fixture("imf_exchangerates.csv")
        next(self.data, None)
        update_exchange_rates(self.data)
        load_fix("transaction-currencies.xml")
        output = list(csv.reader(StringIO(self.client.get(self.base_url).get_data(as_text=True))))
        csv_headers = output[0]
        self.assertEquals(u'11141.99',  output[1][csv_headers.index('total-Commitment-EUR')])
        self.assertEquals(u'6592.92',  output[1][csv_headers.index('total-Disbursement-EUR')])
        self.assertEquals(u'7730.19',  output[1][csv_headers.index('total-Expenditure-EUR')])
        self.assertEquals(u'10004.73',  output[1][csv_headers.index('total-Incoming Funds-EUR')])
        self.assertEquals(u'19102.87',  output[1][csv_headers.index('total-Interest Repayment-EUR')])
        self.assertEquals(u'30475.55',  output[1][csv_headers.index('total-Loan Repayment-EUR')])
        self.assertEquals(u'64593.58',  output[1][csv_headers.index('total-Reimbursement-EUR')])
